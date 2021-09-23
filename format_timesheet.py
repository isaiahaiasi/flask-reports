import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# * Helpers
# todo: replace w builtin openpyxl function
def get_cell(col, row):
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    col_alpha = alpha[col % len(alpha)]
    # TODO: handle columns > Z
    return f"{col_alpha}{row}"


def fmt_time(t_raw):
    try:
        numeric_str = re.sub(r'h|m', '', t_raw) # '3h 15m' -> '3 15'
        str_h, str_m = numeric_str.split(" ")   # '3 15' -> ('3', '15')
        h, m = float(str_h), float(str_m)       # ('3', '15') -> (3.0, 15.0)
        return ((h * 60) + m)/60                # (3.0, 15.0) -> 3.25
    except:
        return None


# * CSV -> DATAFRAME
def get_grouped_dfs(input_file):
    df_raw = pd.read_csv(input_file)

    # split by employee and format xlsx fragment for each person
    dfdict_group = {x: y for x, y in df_raw.groupby('Full Name')}

    for df in dfdict_group.values():
        df.reset_index(inplace=True, drop=True)

    return dfdict_group

def get_col_index(df, col_name):
    return list(df.columns).index(col_name)


# * dict<DATAFRAME> -> XLSX Fns
# write individual timesheet:
# - COLS:
#   - "Hours worked": translation of og Hours Worked/Worked Hours (?) to number format
#   - "UNPAID": Hours worked where UNPAID
#   - "SICK": Empty
#   - "PTO": Empty
#   - "HOLIDAY": Empty
# - SINGLE CELL ENTRIES:
#   - Overtime
# - FORMULAS:
#   - All added cols need SUMs: Hours Worked, OT, SICK, PTO, HOLIDAY
#   - UNPAID/"Hours worked WHERE Break Type=='Unpaid'": SUM (needed to subtract from Hours Worked for REG hours)
#   - REG: Hours worked, minus sum of UNPAID, and minus OT
#   - TOTAL: Hours worked, minus UNPAID, plus SUM of sick, pto, holiday

def get_truncated_df(df, upto_col):
    return df.loc[:, :upto_col]


def write_cell_rows(ws, row, col, vals):
    for i, val in enumerate(vals):
        ws[get_cell(col + i, row)] = val


def add_col_sums(ws, df, col_names, row_start):
    sum_cells = {}
    c_len = len(df.index)

    for col_name in col_names:
        c_index = get_col_index(df, col_name)
        row_end = row_start + c_len

        form_cell = get_cell(c_index, row_start + c_len + 1)
        start_cell = get_cell(c_index, row_start)
        end_cell = get_cell(c_index, row_end)

        ws[form_cell] = f"=SUM({start_cell}:{end_cell})"
        sum_cells[col_name] = form_cell

    return sum_cells


def set_unpaid(df):
    df["UNPAID"] = np.nan
    for i in range(len(df.index)):
        unpaid = df.loc[i, "Break Type"]
        # hrs_worked = float(df.loc[i, "Hours incl break"])
        # unpaid_hrs = hrs_worked if str(unpaid) == "Unpaid" else 0
        # cell_val = unpaid_hrs if unpaid_hrs > 0 else 0
        row_offset = 3
        break_type_cell = get_cell(get_col_index(df, "Break Type"), i + row_offset)
        hrs_worked_cell = get_cell(get_col_index(df, "Hours incl break"), i + row_offset)
        form = f"=IF({break_type_cell}='Unpaid',{hrs_worked_cell},0)"
        df.loc[i, ["UNPAID"]] = form


def write_individual_timesheet(workbook, name, raw_df):
    worksheet = workbook.create_sheet(name)
    worksheet[get_cell(0, 1)] = name  # First row is just employee name

    # grab everything up to x column
    df = get_truncated_df(raw_df, "Break Type")

    # add formatted timesheet entries column
    df["Hours incl break"] = df['Duration'].map(fmt_time)

    set_unpaid(df)

    fillin_cols = ["SICK", "PTO", "HOLIDAY"]

    for col in fillin_cols:
        df[col] = np.nan

    # write contents of dataframe, including headers
    for r_offset in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r_offset)

    r_offset = len(df.index) + 3
    worksheet[get_cell(0, r_offset)] = 'Totals:'

    sum_cols = ["Hours incl break", "UNPAID", *fillin_cols]

    sum_cells = add_col_sums(worksheet, df, sum_cols, 2)

    r_offset = r_offset + 1
    c_offset = list(df.columns).index("Hours incl break") - 1

    # overtime
    worksheet[get_cell(c_offset, r_offset)] = "Overtime:"
    ot_cell = get_cell(c_offset + 1, r_offset)

    r_offset = r_offset + 1

    # REG
    hrs_sum_cell, unpaid_sum_cell = sum_cells['Hours incl break'], sum_cells['UNPAID']
    reg_cell = get_cell(c_offset + 1, r_offset)
    worksheet[get_cell(c_offset, r_offset)] = "REG:"
    worksheet[reg_cell] = f"={hrs_sum_cell}-{unpaid_sum_cell}-{ot_cell}"

    r_offset = r_offset + 1

    # todo: write "grand total" underneath other totals
    sick_sum_c, pto_sum_c, hol_sum_c  = sum_cells["SICK"], sum_cells["PTO"], sum_cells["HOLIDAY"]
    worksheet[get_cell(c_offset, r_offset)] = "TOTAL HRS:"
    tot_form =  f"={reg_cell} + {ot_cell} + {sick_sum_c} + {pto_sum_c} + {hol_sum_c}"
    worksheet[get_cell(c_offset + 1, r_offset)] = tot_form

    print(f"wrote {name}")


def get_xlsx_from_df_group(df_group):
    wb = Workbook()
    for [name, df] in df_group.items():
        write_individual_timesheet(wb, name, df)

    # delete default sheet
    wb.remove(wb["Sheet"])

    return wb


# takes raw csv input_file and returns formatted xlsx file
def get_formatted_timesheet(input_file):
    df_group = get_grouped_dfs(input_file)
    return get_xlsx_from_df_group(df_group)
